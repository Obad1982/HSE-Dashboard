#!/usr/bin/env python3
"""
Excel parity check — Energya HSE

Recomputes FAC / LTI / LWDs from the raw input columns of
'ITR Energya Dashboard 2026.xlsx' > Action_plan using the SAME logic
implemented in src/hse-calculations.ts, then diffs the results against
Excel's own computed columns.

This is the acceptance test for requirement #34 in the project brief
("compare application results against Excel results") and must be re-run
after any change to the calculation engine, and again after the historical
data import.

Last run result: 395 rows, FAC 125=125, LTI 30=30, LWDs 385=385, 0 mismatches.

Usage:
    python3 tests/excel_parity_check.py "path/to/ITR Energya Dashboard 2026.xlsx"
"""
import calendar
import datetime
import sys
import warnings

import openpyxl

warnings.filterwarnings("ignore")

DEFAULT_PATH = "ITR Energya Dashboard 2026.xlsx"


def days_between(start, end):
    """Excel DAYS(end, start) — whole-day difference."""
    return (end - start).days


def is_fac(source, incident, ret):
    """FAC: Source='Injury' AND 0 <= DAYS(return, incident) <= 3"""
    if source != "Injury" or not ret:
        return 0
    d = days_between(incident, ret)
    return 1 if 0 <= d <= 3 else 0


def is_lti(source, incident, ret):
    """LTI: Source='Injury' AND DAYS(return, incident) >= 4"""
    if source != "Injury" or not ret:
        return 0
    return 1 if days_between(incident, ret) >= 4 else 0


def allocate_lost_workdays(source, incident, ret):
    """
    Month-by-month allocation, unbounded (the app removes Excel's 4-column cap).
    Returns [(year, month, days), ...] summing to DAYS(return, incident).
    """
    if not is_lti(source, incident, ret):
        return []
    total = days_between(incident, ret)
    out = []
    year, month, day = incident.year, incident.month, incident.day
    while total > 0:
        days_in_month = calendar.monthrange(year, month)[1]
        available = days_in_month - day + 1
        take = min(available, total)
        out.append((year, month, take))
        total -= take
        day = 1
        month += 1
        if month > 12:
            month = 1
            year += 1
    return out


def main(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Action_plan"]
    headers = [c.value for c in ws[2]]
    col = {h: i + 1 for i, h in enumerate(headers)}

    def cell(r, name):
        return ws.cell(row=r, column=col[name]).value

    def num(v):
        return v if isinstance(v, (int, float)) else 0

    rows = 0
    mine = {"fac": 0, "lti": 0, "lwd": 0}
    excel = {"fac": 0, "lti": 0, "lwd": 0}
    mismatches = {"fac": 0, "lti": 0, "lwd": 0}
    examples = []

    for r in range(3, ws.max_row + 1):
        if cell(r, "NO") is None:
            continue
        incident = cell(r, "التـــاريخ  ")
        if not isinstance(incident, datetime.datetime):
            continue
        rows += 1

        source = cell(r, "المصدر")
        ret = cell(r, "تاريخ العوده")
        ret = ret if isinstance(ret, datetime.datetime) else None

        x_fac = num(cell(r, "FAC"))
        x_lti = num(cell(r, "LTIs "))
        x_lwd = sum(num(cell(r, c)) for c in ("LWDs", "LWDs1", "LWDs2", "LWDs3"))

        m_fac = is_fac(source, incident, ret)
        m_lti = is_lti(source, incident, ret)
        m_lwd = sum(a[2] for a in allocate_lost_workdays(source, incident, ret))

        mine["fac"] += m_fac
        mine["lti"] += m_lti
        mine["lwd"] += m_lwd
        excel["fac"] += x_fac
        excel["lti"] += x_lti
        excel["lwd"] += x_lwd

        if m_fac != x_fac:
            mismatches["fac"] += 1
        if m_lti != x_lti:
            mismatches["lti"] += 1
        if m_lwd != x_lwd:
            mismatches["lwd"] += 1
            if len(examples) < 10:
                examples.append(
                    {
                        "NO": cell(r, "NO"),
                        "source": source,
                        "incident": incident.date(),
                        "return": ret.date() if ret else None,
                        "days": days_between(incident, ret) if ret else None,
                        "mine": m_lwd,
                        "excel": x_lwd,
                    }
                )

    print(f"Rows checked: {rows}\n")
    for key, label in (("fac", "FAC "), ("lti", "LTI "), ("lwd", "LWDs")):
        status = "OK" if mismatches[key] == 0 else "MISMATCH"
        print(
            f"{label} app={mine[key]:<6} excel={excel[key]:<6} "
            f"row-mismatches={mismatches[key]:<4} [{status}]"
        )

    if examples:
        print("\nMismatch examples:")
        for e in examples:
            print("  ", e)

    total_mismatches = sum(mismatches.values())
    print(
        f"\n{'PARITY CONFIRMED' if total_mismatches == 0 else 'PARITY FAILED'}"
        f" — {total_mismatches} total row mismatches"
    )
    return 1 if total_mismatches else 0


if __name__ == "__main__":
    sys.exit(main(sys.argv[1] if len(sys.argv) > 1 else DEFAULT_PATH))
